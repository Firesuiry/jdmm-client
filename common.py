import hashlib
import os
import re
import subprocess
import uuid

import openpyxl

# END_POINT = R'http://api.jiandanmaimai.cn'
END_POINT = R'https://www.jiandanmaimai.cn'
HEAD_STRUCT = bytes('128sIq32s', encoding='utf-8')
FILE_UPLOAD_URL = f'{END_POINT}/file/api/files/'
CLIENT_INFO_URL = f'{END_POINT}/file/api/client/'
FILE_QUERY_URL = f'{END_POINT}/file/api/files/?self=True&limit=20'


url_regex = re.compile(
    r'^(?:http|ftp)s?://'  # http:// or https://
    r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+(?:[A-Z]{2,6}\.?|[A-Z0-9-]{2,}\.?)|'  # domain...
    r'localhost|'  # localhost...
    r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})'  # ...or ip
    r'(?::\d+)?'  # optional port
    r'(?:/?|[/?]\S+)$', re.IGNORECASE)


def check_url(url):
    if url == 'PREPUB':
        return True
    return re.match(url_regex, url) is not None


def cal_md5(file_path):
    with open(file_path, 'rb') as fr:
        md5 = hashlib.md5()
        md5.update(fr.read())
        md5 = md5.hexdigest()
        return md5


def get_file_info(file_path):
    file_name = os.path.basename(file_path)
    file_name_len = len(file_name)
    file_size = os.path.getsize(file_path)
    md5 = cal_md5(file_path)
    return file_name, file_name_len, file_size, md5


def get_Local_ipv6_address():
    child = subprocess.Popen("ipconfig", shell=True, stdout=subprocess.PIPE)
    out = child.communicate();  # 保存ipconfig中的所有信息

    ipv6_pattern = '(([a-f0-9]{1,4}:){7}[a-f0-9]{1,4})'
    try:
        m = re.findall(ipv6_pattern, str(out))
        address = m[1][0]
        print('获取到IPV6地址：{}'.format(address))
        return address
    except Exception as e:
        print(e)
        return


def get_MAC():
    mac = uuid.uuid1().hex[-12:].upper()
    print('当前主机MAC地址为：{}'.format(mac))
    return mac


def write_xlsx(titles, datas, file_path, add=False):
    add = add and os.path.exists(file_path)
    if add:
        book = openpyxl.load_workbook(file_path)
        sheet = book['files']
    else:
        book = openpyxl.Workbook()
        sheet = book.create_sheet('files', 0)

    col = 1
    for title in titles:
        sheet.cell(row=1, column=col).value = title
        col += 1

    row = 2
    have_add_list = []
    if add:
        while True:
            d = sheet.cell(row=row, column=1).value
            if d:
                have_add_list.append(d)
                row += 1
            else:
                break

    for row_data in datas:
        if row_data[0] not in have_add_list:
            col = 1
            for data in row_data:
                sheet.cell(row=row, column=col).value = data
                col += 1
            row += 1
    book.save(file_path)

def has_chinese(s):
    """
    检查整个字符串是否包含中文
    :param s: 需要检查的字符串
    :return: bool
    """
    for ch in s:
        if u'\u4e00' <= ch <= u'\u9fff':
            return True
    return False
