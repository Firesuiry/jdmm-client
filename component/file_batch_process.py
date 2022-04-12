from PySide2.QtCore import QThread, Signal
from PySide2.QtWidgets import QFileDialog
import pathlib
from common import *
import requests

from pathlib import Path
import time
import traceback

path = Path('./data')
FILE_XLS_FILE = 'A.xlsx'

from openpyxl import reader


class ReadXlsxWorker(QThread):
    log_signal = Signal(str)
    set_file_indexs_signal = Signal(list)

    def __init__(self):
        super().__init__()

    def run(self):
        self.log_signal.emit('启动生成所有文件信息列表')
        if not os.path.exists(FILE_XLS_FILE):
            self.log_signal.emit(f'请手动创建文件：{FILE_XLS_FILE}')
            return
        try:
            ws = openpyxl.load_workbook(FILE_XLS_FILE, read_only=True)
            sheet1 = ws.worksheets[0]
            row = 0
            file_indexs = []
            while True:
                row += 1
                file_index = sheet1.cell(row=row, column=1).value
                if file_index is None:
                    break
                if not str(file_index).isdigit():
                    continue
                file_indexs.append(file_index)
            self.set_file_indexs_signal.emit(file_indexs)
        except Exception as e:
            traceback.print_exc()
            self.log_signal.emit(f'出现错误 当前错误：{e}')


class FileBatchProcessWorker(QThread):
    log_signal = Signal(str)
    finish_signal = Signal()

    def __init__(self, cookie, params):
        super().__init__()
        self.cookie = cookie
        self.params = params
        self.indexs = params['indexs']

    def process(self):
        action = self.params['action']
        if len(self.indexs) < 1:
            self.log_signal.emit(f'文件列表为空 处理失败')
            return
        if action == 'money_change':
            self.money_change(self.params['money'])
        elif action == 'money_percent_change':
            self.money_percent_change(self.params['percent'])
        elif action == 'file_off':
            self.file_state_change(False)
        elif action == 'file_on':
            self.file_state_change(True)
        elif action == 'file_add_collection':
            self.file_add_collection(self.params['collection'])
        self.finish_signal.emit()

    def money_change(self, money):
        for index in self.indexs:
            url = f'{END_POINT}/file/api/file/{index}/'
            data = {
                'money': money
            }
            res = requests.post(url, data, cookies=self.cookie)
            time.sleep(1)
            self.log_signal.emit(F"文件：{index}结果：{res.json()['msg']}")

    def money_percent_change(self, percent):
        for index in self.indexs:
            url = f'{END_POINT}/file/api/file/{index}/'
            res = requests.get(url)
            time.sleep(1)
            money = round(res.json()['money'] * percent / 100)
            data = {
                'money': money
            }
            res = requests.post(url, data, cookies=self.cookie)
            time.sleep(1)
            self.log_signal.emit(F"文件：{index}修改至：{money / 100}结果：{res.json()['msg']}")

    def file_state_change(self, on):
        for index in self.indexs:
            url = f'{END_POINT}/file/api/file/{index}/'
            data = {
                'state': 'checking' if on else 'off'
            }
            res = requests.post(url, data, cookies=self.cookie)
            time.sleep(1)
            self.log_signal.emit(F"文件：{index}结果：{res.json()['msg']}")

    def file_add_collection(self, collection):
        for index in self.indexs:
            url = f'{END_POINT}/file/api/collectionop/{collection}/'
            data = {
                'action': 'add_file',
                'file': index
            }
            res = requests.post(url, data, cookies=self.cookie)
            time.sleep(1)
            self.log_signal.emit(F"文件：{index}结果：{res.json()['msg']}")

    def run(self):
        try:
            self.process()
        except Exception as e:
            print(e)
            self.log_signal.emit(f'出现错误 当前错误：{e}')


class FileBatchManager:
    def __init__(self, window, main):
        self.window = window
        self.main = main
        self.add_log('欢迎使用账号管理工具')
        self.window.file_list_read_btn.clicked.connect(self.on_file_list_read_btn_clicked)
        self.window.file_money_change_btn.clicked.connect(self.on_file_money_change_btn_clicked)
        self.window.file_money_precent_change_btn.clicked.connect(self.on_file_money_precent_change_btn_clicked)
        self.window.file_off_btn.clicked.connect(self.on_file_off_btn_clicked)
        self.window.file_on_btn.clicked.connect(self.on_file_on_btn_clicked)
        self.window.file_add_collection_btn.clicked.connect(self.on_file_add_collection_btn_clicked)
        # self.window.generate_file_list_btn.clicked.connect(self.on_generate_file_list_btn_clicked)
        self.running = False

    def on_generate_file_list_btn_clicked(self):
        book = openpyxl.Workbook()
        # sheet = book.create_sheet('files', 0)
        book.save(FILE_XLS_FILE)
        self.add_log('生成文件列表成功')

    def get_file_list(self):
        s = self.window.file_list.toPlainText()
        # print(s)
        strs = s.split()
        indexs = []
        for s in strs:
            if str(s).isdigit():
                indexs.append(int(s))
        print(indexs)
        return indexs

    def on_file_list_read_btn_clicked(self):
        self.add_log('正在读取文件列表')

        self.generate_file_xls_worker = ReadXlsxWorker()
        self.generate_file_xls_worker.log_signal.connect(self.add_log)
        self.generate_file_xls_worker.set_file_indexs_signal.connect(self.set_file_indexs)
        self.generate_file_xls_worker.start()

    def on_file_money_change_btn_clicked(self):
        money = round(100 * float(self.window.money_edit.text()))
        self.add_log(F'正在修改文件价格至：{money / 100}元')
        if not self.run_thread_flag():
            self.add_log('有进程正在执行')
            return
        params = {
            'indexs': self.get_file_list(),
            'action': 'money_change',
            'money': money,
        }
        self.file_batch_process_worker = FileBatchProcessWorker(self.main.cookie, params)
        self.file_batch_process_worker.log_signal.connect(self.add_log)
        self.file_batch_process_worker.finish_signal.connect(self.process_finish)

        self.file_batch_process_worker.start()

    def on_file_money_precent_change_btn_clicked(self):
        percent = int(self.window.money_percent_edit.text())
        self.add_log(F'正在修改文件价格至：{percent}%')
        if not self.run_thread_flag():
            self.add_log('有进程正在执行')
            return
        if percent > 999:
            self.add_log('百分比不能超过999')
            return
        if percent < 10:
            self.add_log('百分比不能小于10')
            return
        params = {
            'indexs': self.get_file_list(),
            'action': 'money_percent_change',
            'percent': percent,
        }
        self.file_batch_process_worker = FileBatchProcessWorker(self.main.cookie, params)
        self.file_batch_process_worker.log_signal.connect(self.add_log)
        self.file_batch_process_worker.finish_signal.connect(self.process_finish)

        self.file_batch_process_worker.start()

    def on_file_off_btn_clicked(self):
        self.add_log(F'正在下架文件')
        if not self.run_thread_flag():
            self.add_log('有进程正在执行')
            return
        params = {
            'indexs': self.get_file_list(),
            'action': 'file_off',
        }
        self.file_batch_process_worker = FileBatchProcessWorker(self.main.cookie, params)
        self.file_batch_process_worker.log_signal.connect(self.add_log)
        self.file_batch_process_worker.finish_signal.connect(self.process_finish)

        self.file_batch_process_worker.start()

    def on_file_on_btn_clicked(self):
        self.add_log(F'正在上架文件')
        if not self.run_thread_flag():
            self.add_log('有进程正在执行')
            return
        params = {
            'indexs': self.get_file_list(),
            'action': 'file_on',
        }
        self.file_batch_process_worker = FileBatchProcessWorker(self.main.cookie, params)
        self.file_batch_process_worker.log_signal.connect(self.add_log)
        self.file_batch_process_worker.finish_signal.connect(self.process_finish)

        self.file_batch_process_worker.start()

    def on_file_add_collection_btn_clicked(self):
        collection = int(self.window.collection_id_edit.text())
        self.add_log(F'正在添加文件至收藏夹：{collection}')
        if not self.run_thread_flag():
            self.add_log('有进程正在执行')
            return
        params = {
            'indexs': self.get_file_list(),
            'action': 'file_add_collection',
            'collection': collection,
        }
        self.file_batch_process_worker = FileBatchProcessWorker(self.main.cookie, params)
        self.file_batch_process_worker.log_signal.connect(self.add_log)
        self.file_batch_process_worker.finish_signal.connect(self.process_finish)
        self.file_batch_process_worker.start()

    def add_log(self, msg):
        print('添加log：{}'.format(msg))
        self.window.file_batch_info.append(str(msg))
        self.window.file_batch_info.moveCursor(self.window.client_log_text.textCursor().End)

    def set_file_indexs(self, file_indexs):

        s = ' '.join([str(i) for i in file_indexs])
        self.add_log('写入文件列表' + s)

        # self.window.file_list.clear()
        self.window.file_list.setPlainText(s)

    def process_finish(self):
        self.add_log('执行完成')
        self.running = False

    def run_thread_flag(self):
        if self.running:
            return False
        self.running = True
        return True
